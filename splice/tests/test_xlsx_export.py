"""
Tests for Excel workbook export.
"""

import os

import numpy as np
import pytest

from splice.core.diff import DiffResult
from splice.core.diagnostics import EventDiagnostic
from splice.core.junction_extractor import JunctionEvidence
from splice.io.xlsx_export import export_xlsx_workbook
from splice.utils.genomic import Junction


def _make_diff_result(module_id, gene_name, event_type, p_value, fdr, delta=0.3):
    return DiffResult(
        module_id=module_id, gene_id=f"G_{module_id}", gene_name=gene_name,
        chrom="chr1", strand="+", event_type=event_type, n_junctions=2,
        junction_coords=["chr1:100-200:+", "chr1:300-400:+"],
        junction_confidence=[0.9, 0.8], is_annotated=[True, False],
        psi_group1=np.array([0.8, 0.2]), psi_group2=np.array([0.5, 0.5]),
        delta_psi=np.array([-delta, delta]), max_abs_delta_psi=delta,
        delta_psi_ci_low=np.array([-0.4, 0.1]),
        delta_psi_ci_high=np.array([-0.2, 0.4]),
        log_likelihood_null=-100.0, log_likelihood_full=-90.0,
        degrees_of_freedom=1, p_value=p_value, fdr=fdr,
        null_converged=True, full_converged=True, null_refit_used=False,
        null_iterations=10, full_iterations=10,
        null_gradient_norm=0.01, full_gradient_norm=0.01,
    )


def _make_diagnostic(module_id, tier="HIGH"):
    return EventDiagnostic(
        module_id=module_id, confidence_tier=tier,
        null_converged=True, full_converged=True, null_refit_used=False,
        mean_mapq=55.0, median_mapq=55.0,
        frac_high_mapq=0.9, frac_multi_mapped=0.05,
        min_group_total_reads=50, effective_n_min=3.0,
        mean_junction_confidence=0.85, min_junction_confidence=0.7,
        frac_annotated_junctions=0.8,
        has_novel_junctions=True, has_low_confidence_junction=False,
        has_convergence_issue=False, reason="All criteria met",
        bootstrap_cv=0.05, prior_dominance=0.1,
    )


class TestXlsxExport:

    def test_creates_file(self, tmp_path):
        """XLSX file should be created."""
        diff = [_make_diff_result("m1", "GeneA", "SE", 0.001, 0.01)]
        diag = [_make_diagnostic("m1")]
        path = str(tmp_path / "test.xlsx")

        export_xlsx_workbook(diff, diag, {}, None, {"SE": 1}, path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    def test_has_all_sheets(self, tmp_path):
        """Workbook should have all 4 sheets."""
        from openpyxl import load_workbook

        diff = [_make_diff_result("m1", "GeneA", "SE", 0.001, 0.01)]
        diag = [_make_diagnostic("m1")]
        path = str(tmp_path / "test.xlsx")

        export_xlsx_workbook(diff, diag, {}, None, {"SE": 1}, path)

        wb = load_workbook(path)
        assert "Results" in wb.sheetnames
        assert "Junction Details" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Event Types" in wb.sheetnames

    def test_results_row_count(self, tmp_path):
        """Results sheet should have header + one row per DiffResult."""
        from openpyxl import load_workbook

        diff = [
            _make_diff_result("m1", "GeneA", "SE", 0.001, 0.01),
            _make_diff_result("m2", "GeneB", "A3SS", 0.05, 0.08),
            _make_diff_result("m3", "GeneC", "Complex", 0.2, 0.5),
        ]
        diag = [_make_diagnostic(f"m{i+1}") for i in range(3)]
        path = str(tmp_path / "test.xlsx")

        export_xlsx_workbook(diff, diag, {}, None, {"SE": 1, "A3SS": 1, "Complex": 1}, path)

        wb = load_workbook(path)
        ws = wb["Results"]
        assert ws.max_row == 4  # 1 header + 3 data rows

    def test_empty_results(self, tmp_path):
        """Should handle empty results without error."""
        path = str(tmp_path / "test.xlsx")
        export_xlsx_workbook([], [], {}, None, {}, path)
        assert os.path.exists(path)

    def test_junction_details_with_evidence(self, tmp_path):
        """Junction Details sheet should contain junction data."""
        from openpyxl import load_workbook

        j1 = Junction("chr1", 100, 200, "+")
        ev1 = JunctionEvidence(
            junction=j1,
            sample_counts=np.array([10, 5]),
            sample_weighted_counts=np.array([10.0, 5.0]),
            sample_mapq_mean=np.array([60.0, 55.0]),
            sample_mapq_median=np.array([60.0, 55.0]),
            sample_nh_distribution=np.array([1.0, 1.0]),
            is_annotated=True, motif="GT/AG", motif_score=1.0,
            max_anchor=50, n_samples_detected=2, cross_sample_recurrence=1.0,
        )

        path = str(tmp_path / "test.xlsx")
        export_xlsx_workbook([], [], {j1: ev1}, None, {}, path)

        wb = load_workbook(path)
        ws = wb["Junction Details"]
        assert ws.max_row == 2  # header + 1 junction
