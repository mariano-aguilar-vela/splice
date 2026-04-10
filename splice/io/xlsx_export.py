"""
Publication-quality Excel workbook export for SPLICE results.

Creates a multi-sheet Excel workbook with conditional formatting,
auto-column-widths, and structured data for downstream analysis.
"""

from __future__ import annotations

from typing import Dict, List, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from splice.core.diff import DiffResult
from splice.core.diagnostics import EventDiagnostic


def _auto_width(ws, min_width=8, max_width=40):
    """Set column widths based on content."""
    for col in ws.columns:
        lengths = []
        for cell in col:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _header_style():
    """Return header cell formatting."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin"),
            right=Side(style="thin"),
        ),
    }


def _apply_header(ws, row=1):
    """Apply header styling to the first row."""
    style = _header_style()
    for cell in ws[row]:
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]


def _write_results_sheet(wb, diff_results, diagnostics):
    """Write the Results sheet with conditional formatting."""
    ws = wb.active
    ws.title = "Results"

    # Build diagnostics lookup
    diag_map = {}
    if diagnostics:
        for d in diagnostics:
            diag_map[d.module_id] = d

    # Headers
    headers = [
        "module_id", "gene_id", "gene_name", "chrom", "strand",
        "event_type", "n_junctions", "max_abs_delta_psi",
        "p_value", "fdr", "confidence_tier",
        "null_converged", "full_converged", "null_refit_used",
    ]
    ws.append(headers)
    _apply_header(ws)

    # Conditional formatting fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for dr in diff_results:
        diag = diag_map.get(dr.module_id)
        confidence_tier = diag.confidence_tier if diag else "NA"

        row = [
            dr.module_id,
            dr.gene_id,
            dr.gene_name,
            dr.chrom,
            dr.strand,
            dr.event_type,
            dr.n_junctions,
            f"{dr.max_abs_delta_psi:.4f}",
            f"{dr.p_value:.6e}",
            f"{dr.fdr:.6e}",
            confidence_tier,
            str(dr.null_converged),
            str(dr.full_converged),
            str(dr.null_refit_used),
        ]
        ws.append(row)

        # Conditional formatting on FDR
        row_idx = ws.max_row
        fdr_cell = ws.cell(row=row_idx, column=10)
        if dr.fdr < 0.05:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = green_fill
        elif dr.fdr < 0.1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = yellow_fill

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_junction_details_sheet(wb, junction_evidence, confidence_scores):
    """Write the Junction Details sheet."""
    ws = wb.create_sheet("Junction Details")

    headers = [
        "junction_id", "chrom", "start", "end", "strand",
        "is_annotated", "motif", "motif_score", "confidence_score",
        "total_reads", "n_samples_detected", "cross_sample_recurrence",
    ]
    ws.append(headers)
    _apply_header(ws)

    for junc, ev in junction_evidence.items():
        junc_id = f"{junc.chrom}:{junc.start}-{junc.end}:{junc.strand}"

        conf_score = 0.0
        if confidence_scores and junc in confidence_scores:
            conf = confidence_scores[junc]
            conf_score = conf.composite_score if hasattr(conf, 'composite_score') else float(conf)

        row = [
            junc_id,
            junc.chrom,
            junc.start,
            junc.end,
            junc.strand,
            str(ev.is_annotated),
            ev.motif,
            f"{ev.motif_score:.4f}",
            f"{conf_score:.4f}",
            int(np.sum(ev.sample_counts)),
            ev.n_samples_detected,
            f"{ev.cross_sample_recurrence:.4f}",
        ]
        ws.append(row)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb, diff_results, diagnostics, event_type_counts):
    """Write the Summary sheet."""
    ws = wb.create_sheet("Summary")

    bold = Font(bold=True, size=11)

    def _add_section(title, rows):
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
        ws.append([])
        for label, value in rows:
            ws.append([label, value])
            ws.cell(row=ws.max_row, column=1).font = bold
        ws.append([])

    # Significance counts
    n_total = len(diff_results)
    n_sig_005 = sum(1 for dr in diff_results if dr.fdr < 0.05)
    n_sig_001 = sum(1 for dr in diff_results if dr.fdr < 0.01)

    _add_section("Significance", [
        ("Total modules tested", n_total),
        ("Significant (FDR < 0.05)", n_sig_005),
        ("Significant (FDR < 0.01)", n_sig_001),
        ("Significance rate (FDR < 0.05)", f"{n_sig_005/n_total*100:.1f}%" if n_total > 0 else "0%"),
    ])

    # Event types
    _add_section("Event Types", [
        (evt, count) for evt, count in sorted(event_type_counts.items())
    ])

    # Convergence
    n_null_conv = sum(1 for dr in diff_results if dr.null_converged)
    n_full_conv = sum(1 for dr in diff_results if dr.full_converged)
    n_refit = sum(1 for dr in diff_results if dr.null_refit_used)

    _add_section("Convergence", [
        ("Null model converged", f"{n_null_conv}/{n_total}"),
        ("Full model converged", f"{n_full_conv}/{n_total}"),
        ("Null-refit applied", f"{n_refit}/{n_total}"),
    ])

    # Confidence tiers
    if diagnostics:
        tier_counts = {}
        for d in diagnostics:
            tier_counts[d.confidence_tier] = tier_counts.get(d.confidence_tier, 0) + 1
        _add_section("Confidence Tiers", [
            (tier, count) for tier, count in sorted(tier_counts.items())
        ])

        # Mean MAPQ
        mapqs = [d.mean_mapq for d in diagnostics if d.mean_mapq > 0]
        if mapqs:
            _add_section("Quality Metrics", [
                ("Mean MAPQ (across events)", f"{np.mean(mapqs):.1f}"),
                ("Median MAPQ (across events)", f"{np.median(mapqs):.1f}"),
            ])

    _auto_width(ws)


def _write_event_types_sheet(wb, diff_results, event_type_counts):
    """Write the Event Types breakdown sheet."""
    ws = wb.create_sheet("Event Types")

    headers = ["Event Type", "Total Count", "Significant (FDR<0.05)", "Percentage of Total"]
    ws.append(headers)
    _apply_header(ws)

    n_total = len(diff_results)

    # Count significant per type
    sig_by_type = {}
    for dr in diff_results:
        if dr.fdr < 0.05:
            sig_by_type[dr.event_type] = sig_by_type.get(dr.event_type, 0) + 1

    for evt in sorted(event_type_counts.keys()):
        count = event_type_counts[evt]
        sig_count = sig_by_type.get(evt, 0)
        pct = f"{count/n_total*100:.1f}%" if n_total > 0 else "0%"
        ws.append([evt, count, sig_count, pct])

    _auto_width(ws)
    ws.freeze_panes = "A2"


def export_xlsx_workbook(
    diff_results: List[DiffResult],
    diagnostics: List[EventDiagnostic],
    junction_evidence: Dict,
    confidence_scores: Optional[Dict],
    event_type_counts: Dict[str, int],
    output_path: str,
) -> None:
    """Create a publication-quality Excel workbook with SPLICE results.

    Args:
        diff_results: List of DiffResult objects.
        diagnostics: List of EventDiagnostic objects.
        junction_evidence: Dict mapping Junction to JunctionEvidence.
        confidence_scores: Dict mapping Junction to confidence scores.
        event_type_counts: Dict mapping event type to count.
        output_path: Path to output .xlsx file.
    """
    wb = Workbook()

    _write_results_sheet(wb, diff_results, diagnostics)
    _write_junction_details_sheet(wb, junction_evidence, confidence_scores)
    _write_summary_sheet(wb, diff_results, diagnostics, event_type_counts)
    _write_event_types_sheet(wb, diff_results, event_type_counts)

    wb.save(output_path)
